from openpyxl import Workbook

from popcorn.structures import Event

def _hotspots_sheet_name(item_name: str) -> str:
    return str("pops__"+item_name)

def _kernel_differences_sheet_name(item_name: str) -> str:
    return str("kdiffs__"+item_name)

def report_hotspots(result: dict[str, list[Event]], wb: Workbook):
    items = list(result.items())
    count = len(items)
    if wb.active.title == "Sheet": # current sheet is unused
        ws = wb.active
        ws.title = _hotspots_sheet_name(items[0][0])
    else:
        ws = wb.create_sheet(_hotspots_sheet_name(items[0][0]))

    for i in range(0, count):
        ws.append(['ph','tid','pid','name','cat','ts','id','dur','args_id']) # header
        for event in items[i][1]:
            ws.append([event.ph, event.tid, event.pid, event.name, event.cat, event.ts, event.id, event.dur, event.args_id])
        if i < count - 1:
            ws = wb.create_sheet(_hotspots_sheet_name(items[i+1][0]))

def report_kdiffs(result: dict[str, list[tuple[Event, int]]], wb: Workbook):
    items = list(result.items())
    count = len(items)
    if wb.active.title == "Sheet": # current sheet is unused
        ws = wb.active
        ws.title = _kernel_differences_sheet_name(items[0][0])
    else:
        ws = wb.create_sheet(_kernel_differences_sheet_name(items[0][0]))

    for i in range(0, count):
        ws.append(['diff','ph','tid','pid','name','cat','ts','id','dur','args_id']) # header
        for (event, diff) in items[i][1]:
            ws.append([diff, event.ph, event.tid, event.pid, event.name, event.cat, event.ts, event.id, event.dur, event.args_id])
        if i < count - 1:
            ws = wb.create_sheet(_kernel_differences_sheet_name(items[i+1][0]))