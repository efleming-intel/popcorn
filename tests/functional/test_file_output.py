from openpyxl import load_workbook
import os
import shutil
from unittest.mock import patch

from popcorn.__main__ import main_cli
from popcorn.reporters import _hotspots_sheet_name, _kernel_differences_sheet_name


def test_xlsx_output(request, tmp_path):
    inputdir = str(os.path.dirname(request.fspath)) + "/data/"
    input_file = inputdir + "trace1.json"
    output = str(tmp_path) + "/result"

    testargs = ["popcorn", input_file, "-ot", "xlsx", "-o", output]

    with patch("sys.argv", testargs):
        errors = main_cli()
        assert errors == None

    expected_result_file = os.path.abspath(str(output + ".xlsx"))
    assert os.path.exists(expected_result_file)

    wb = load_workbook(expected_result_file)
    expected_worksheet_name = _hotspots_sheet_name(
        os.path.basename(input_file).removesuffix(".json")
    )
    assert expected_worksheet_name in wb.sheetnames

    os.remove(expected_result_file)


def test_md_output(request, tmp_path):
    inputdir = str(os.path.dirname(request.fspath)) + "/data/"
    input_file = inputdir + "trace1.json"
    output = str(tmp_path) + "/result"

    testargs = ["popcorn", "-ot", "md", "-o", output, input_file]

    with patch("sys.argv", testargs):
        errors = main_cli()
        assert errors == None

    expected_result_file = os.path.abspath(str(output + ".md"))
    assert os.path.exists(expected_result_file)

    with open(expected_result_file, "r") as file:
        contents = "".join(file.readlines())
        expected_worksheet_name = _hotspots_sheet_name(
            os.path.basename(input_file).removesuffix(".json")
        )
        assert expected_worksheet_name in contents

    os.remove(expected_result_file)


def test_csv_output(request, tmp_path):
    inputdir = str(os.path.dirname(request.fspath)) + "/data/"
    input_files = [str(inputdir + "trace1.json"), str(inputdir + "trace0.json")]
    output = str(tmp_path) + "/result"

    testargs = ["popcorn", input_files[0], input_files[1], "-ot", "csv", "-o", output]

    with patch("sys.argv", testargs):
        errors = main_cli()
        assert errors == None

    assert os.path.exists(output)
    expected_result_names = [
        _hotspots_sheet_name(os.path.basename(input_files[0]).removesuffix(".json")),
        _hotspots_sheet_name(os.path.basename(input_files[1]).removesuffix(".json")),
        _kernel_differences_sheet_name(
            f"{os.path.basename(input_files[0]).removesuffix('.json')}_{os.path.basename(input_files[1]).removesuffix('.json')}"
        )
    ]
    expected_result_files = [
        os.path.abspath(f"{output}/{name}.csv") for name in expected_result_names
    ]

    for expected_file in expected_result_files:
        assert os.path.exists(expected_file)

    shutil.rmtree(output)
